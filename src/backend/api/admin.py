"""
后台管理 API - 仪表盘、台账管理、导出、配置
"""
import io
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from models.database import (
    SessionLocal, LawyerDiscipline, LegislationProject, JudicialCase,
    JudicialPolicy, LegalRecruitment, Message, EventRegistration,
    BasicContent, CollectionLog, AdminUser
)
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

admin_bp = Blueprint('admin', __name__)


# ── 简单 token 验证（生产环境应使用 JWT） ──────────────
def verify_admin_token():
    """验证管理员 token，返回 admin_user 或 None"""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    if not token:
        return None
    db = SessionLocal()
    try:
        user = db.query(AdminUser).filter_by(username=token.split(':')[0]).first()
        if user and check_password_hash(user.password_hash, token.split(':')[1] if ':' in token else ''):
            return user
        return None
    finally:
        db.close()


def require_admin():
    user = verify_admin_token()
    if not user:
        return jsonify({'code': 401, 'message': '未授权访问', 'data': None}), 401
    return user


# ══════════════════════════════════════════════════════
# 登录
# ══════════════════════════════════════════════════════
@admin_bp.route('/admin/login', methods=['POST'])
def admin_login():
    data = request.get_json() or {}
    username = data.get('username', '')
    password = data.get('password', '')

    db = SessionLocal()
    try:
        user = db.query(AdminUser).filter_by(username=username, is_active=True).first()
        if not user or not check_password_hash(user.password_hash, password):
            return {'code': 401, 'message': '用户名或密码错误', 'data': None}, 401

        user.last_login = datetime.utcnow()
        db.commit()

        # 简单的 token: username:password_hash
        token = f'{user.username}:{user.password_hash}'
        return {'code': 200, 'message': '登录成功', 'data': {
            'token': token,
            'username': user.username,
            'role': user.role,
        }}
    finally:
        db.close()


# ══════════════════════════════════════════════════════
# 仪表盘
# ══════════════════════════════════════════════════════
@admin_bp.route('/admin/dashboard')
def admin_dashboard():
    user = require_admin()
    if not isinstance(user, AdminUser):
        return user

    db = SessionLocal()
    try:
        now = datetime.utcnow()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

        dashboard = {
            'totals': {
                'discipline': db.query(func.count(LawyerDiscipline.id)).scalar() or 0,
                'legislation': db.query(func.count(LegislationProject.id)).scalar() or 0,
                'cases': db.query(func.count(JudicialCase.id)).scalar() or 0,
                'policy': db.query(func.count(JudicialPolicy.id)).scalar() or 0,
                'recruitment': db.query(func.count(LegalRecruitment.id)).filter(LegalRecruitment.is_active == True).scalar() or 0,
            },
            'today_new': {
                'discipline': db.query(func.count(LawyerDiscipline.id)).filter(LawyerDiscipline.created_at >= today_start).scalar() or 0,
                'legislation': db.query(func.count(LegislationProject.id)).filter(LegislationProject.created_at >= today_start).scalar() or 0,
                'cases': db.query(func.count(JudicialCase.id)).filter(JudicialCase.created_at >= today_start).scalar() or 0,
                'policy': db.query(func.count(JudicialPolicy.id)).filter(JudicialPolicy.created_at >= today_start).scalar() or 0,
                'recruitment': db.query(func.count(LegalRecruitment.id)).filter(LegalRecruitment.created_at >= today_start).scalar() or 0,
            },
            'messages_unread': db.query(func.count(Message.id)).filter(Message.is_read == False).scalar() or 0,
            'registrations_pending': db.query(func.count(EventRegistration.id)).filter(EventRegistration.status == 'pending').scalar() or 0,
            'last_collection': None,
        }

        # 最近采集记录
        last_log = db.query(CollectionLog).order_by(CollectionLog.started_at.desc()).first()
        if last_log:
            dashboard['last_collection'] = {
                'name': last_log.collector_name,
                'status': last_log.status,
                'started_at': last_log.started_at.isoformat(),
                'new_added': last_log.new_added,
            }

        return {'code': 200, 'message': 'success', 'data': dashboard}
    finally:
        db.close()


# ══════════════════════════════════════════════════════
# Excel 导出
# ══════════════════════════════════════════════════════
def _export_to_excel(query, filename, headers, row_mapper):
    """通用 Excel 导出"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '数据'

    # 表头样式
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='203B64', end_color='203B64', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, item in enumerate(query.all(), 2):
        row_data = row_mapper(item)
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # 自动列宽
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@admin_bp.route('/admin/export/discipline')
def export_discipline():
    user = require_admin()
    if not isinstance(user, AdminUser):
        return user

    db = SessionLocal()
    try:
        headers = ['ID', '律师姓名', '律所', '省份', '城市', '处分类型', '处分日期', '处分机构', '违规详情']
        def mapper(i):
            return [i.id, i.lawyer_name, i.law_firm, i.firm_province, i.firm_city,
                    i.discipline_type, i.discipline_date.strftime('%Y-%m-%d') if i.discipline_date else '',
                    i.discipline_org, i.violation_detail]
        return _export_to_excel(db.query(LawyerDiscipline).order_by(LawyerDiscipline.discipline_date.desc()),
                                '律师行政处罚数据.xlsx', headers, mapper)
    finally:
        db.close()


@admin_bp.route('/admin/export/legislation')
def export_legislation():
    user = require_admin()
    if not isinstance(user, AdminUser):
        return user
    db = SessionLocal()
    try:
        headers = ['ID', '标题', '阶段', '类别', '提出机构', '发布日期', '生效日期', '摘要']
        def mapper(i):
            return [i.id, i.title, i.stage, i.category, i.proposing_body,
                    i.draft_date.strftime('%Y-%m-%d') if i.draft_date else '',
                    i.effective_date.strftime('%Y-%m-%d') if i.effective_date else '', i.summary]
        return _export_to_excel(db.query(LegislationProject).order_by(LegislationProject.updated_at.desc()),
                                '立法项目数据.xlsx', headers, mapper)
    finally:
        db.close()


@admin_bp.route('/admin/export/cases')
def export_cases():
    user = require_admin()
    if not isinstance(user, AdminUser):
        return user
    db = SessionLocal()
    try:
        headers = ['ID', '案件标题', '案号', '类型', '案由', '阶段', '法院', '立案日期', '裁判日期']
        def mapper(i):
            return [i.id, i.case_title, i.case_number, i.case_type, i.cause_of_action,
                    i.stage, i.court,
                    i.filing_date.strftime('%Y-%m-%d') if i.filing_date else '',
                    i.ruling_date.strftime('%Y-%m-%d') if i.ruling_date else '']
        return _export_to_excel(db.query(JudicialCase).order_by(JudicialCase.ruling_date.desc()),
                                '司法案件数据.xlsx', headers, mapper)
    finally:
        db.close()


@admin_bp.route('/admin/export/policy')
def export_policy():
    user = require_admin()
    if not isinstance(user, AdminUser):
        return user
    db = SessionLocal()
    try:
        headers = ['ID', '标题', '文号', '类别', '发布机构', '发布日期', '生效日期']
        def mapper(i):
            return [i.id, i.title, i.doc_number, i.category, i.issuing_body,
                    i.publish_date.strftime('%Y-%m-%d') if i.publish_date else '',
                    i.effective_date.strftime('%Y-%m-%d') if i.effective_date else '']
        return _export_to_excel(db.query(JudicialPolicy).order_by(JudicialPolicy.publish_date.desc()),
                                '司法政策数据.xlsx', headers, mapper)
    finally:
        db.close()


@admin_bp.route('/admin/export/recruitment')
def export_recruitment():
    user = require_admin()
    if not isinstance(user, AdminUser):
        return user
    db = SessionLocal()
    try:
        headers = ['ID', '标题', '公司', '城市', '岗位类型', '薪资', '经验', '学历', '发布日期']
        def mapper(i):
            return [i.id, i.title, i.company, i.city, i.job_type, i.salary_range,
                    i.experience, i.education,
                    i.publish_date.strftime('%Y-%m-%d') if i.publish_date else '']
        return _export_to_excel(db.query(LegalRecruitment).filter(LegalRecruitment.is_active == True).order_by(LegalRecruitment.publish_date.desc()),
                                '法律招聘数据.xlsx', headers, mapper)
    finally:
        db.close()


# ══════════════════════════════════════════════════════
# 留言管理
# ══════════════════════════════════════════════════════
@admin_bp.route('/admin/messages')
def admin_messages():
    user = require_admin()
    if not isinstance(user, AdminUser):
        return user
    db = SessionLocal()
    try:
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 20, type=int)
        is_read = request.args.get('is_read', '').strip()

        query = db.query(Message)
        if is_read == '0':
            query = query.filter(Message.is_read == False)
        elif is_read == '1':
            query = query.filter(Message.is_read == True)

        total = query.count()
        items = query.order_by(Message.created_at.desc()).offset((page-1)*page_size).limit(page_size).all()

        return {'code': 200, 'message': 'success', 'data': {
            'items': [{
                'id': m.id, 'name': m.name, 'email': m.email, 'phone': m.phone,
                'company': m.company, 'content': m.content, 'category': m.category,
                'is_read': m.is_read, 'is_replied': m.is_replied,
                'created_at': m.created_at.isoformat(),
            } for m in items],
            'pagination': {'total': total, 'page': page, 'page_size': page_size}
        }}
    finally:
        db.close()


@admin_bp.route('/admin/messages/<int:id>/read', methods=['POST'])
def mark_message_read(id):
    user = require_admin()
    if not isinstance(user, AdminUser):
        return user
    db = SessionLocal()
    try:
        msg = db.query(Message).filter_by(id=id).first()
        if msg:
            msg.is_read = True
            db.commit()
        return {'code': 200, 'message': 'success', 'data': None}
    finally:
        db.close()


# ══════════════════════════════════════════════════════
# 报名管理
# ══════════════════════════════════════════════════════
@admin_bp.route('/admin/registrations')
def admin_registrations():
    user = require_admin()
    if not isinstance(user, AdminUser):
        return user
    db = SessionLocal()
    try:
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 20, type=int)

        query = db.query(EventRegistration)
        total = query.count()
        items = query.order_by(EventRegistration.created_at.desc()).offset((page-1)*page_size).limit(page_size).all()

        return {'code': 200, 'message': 'success', 'data': {
            'items': [{
                'id': r.id, 'event_title': r.event_title, 'name': r.name,
                'phone': r.phone, 'email': r.email, 'company': r.company,
                'position': r.position, 'status': r.status,
                'created_at': r.created_at.isoformat(),
            } for r in items],
            'pagination': {'total': total, 'page': page, 'page_size': page_size}
        }}
    finally:
        db.close()


# ══════════════════════════════════════════════════════
# 基础内容管理
# ══════════════════════════════════════════════════════
@admin_bp.route('/admin/content', methods=['GET', 'POST'])
def admin_content():
    user = require_admin()
    if not isinstance(user, AdminUser):
        return user
    db = SessionLocal()
    try:
        if request.method == 'GET':
            items = db.query(BasicContent).order_by(BasicContent.order_index).all()
            return {'code': 200, 'message': 'success', 'data': [{
                'id': c.id, 'content_key': c.content_key, 'title': c.title,
                'content_value': c.content_value, 'is_published': c.is_published,
            } for c in items]}

        # POST: 更新内容
        data = request.get_json() or {}
        key = data.get('content_key')
        item = db.query(BasicContent).filter_by(content_key=key).first()
        if item:
            item.content_value = data.get('content_value', item.content_value)
            item.title = data.get('title', item.title)
            item.is_published = data.get('is_published', item.is_published)
        else:
            item = BasicContent(
                content_key=key,
                title=data.get('title', ''),
                content_value=data.get('content_value', ''),
            )
            db.add(item)
        db.commit()
        return {'code': 200, 'message': '保存成功', 'data': None}
    finally:
        db.close()


# ══════════════════════════════════════════════════════
# 采集日志
# ══════════════════════════════════════════════════════
@admin_bp.route('/admin/collection-logs')
def collection_logs():
    user = require_admin()
    if not isinstance(user, AdminUser):
        return user
    db = SessionLocal()
    try:
        logs = db.query(CollectionLog).order_by(CollectionLog.started_at.desc()).limit(50).all()
        return {'code': 200, 'message': 'success', 'data': [{
            'id': l.id, 'collector_name': l.collector_name, 'status': l.status,
            'total_fetched': l.total_fetched, 'new_added': l.new_added,
            'duplicates': l.duplicates, 'errors': l.errors,
            'started_at': l.started_at.isoformat(),
            'finished_at': l.finished_at.isoformat() if l.finished_at else '',
            'duration_seconds': l.duration_seconds,
        } for l in logs]}
    finally:
        db.close()
